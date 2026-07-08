"""RAGAS 평가 서비스"""

import asyncio
import io
import logging
import math
from typing import Any, BinaryIO, Dict, List, Optional

import openpyxl
from openai import AsyncOpenAI
from ragas import evaluate
from ragas.dataset_schema import EvaluationDataset, SingleTurnSample
from ragas.embeddings.base import embedding_factory
from ragas.llms import llm_factory
from ragas.metrics._answer_relevance import AnswerRelevancy
from ragas.metrics._context_precision import ContextPrecision
from ragas.metrics._context_recall import ContextRecall
from ragas.metrics._faithfulness import Faithfulness
from ragas.run_config import RunConfig

from app.dto.ragas_dto import RagasDatasetItem
from app.utils.llm_compat import needs_max_completion_tokens

logger = logging.getLogger(__name__)

REQUIRED_COLUMNS = {
    "id", "user_input", "category", "reference_context_1", "source_document",
    "reference_page_1", "reference_page_2", "reference_page_3",
}


def _validate_page_value(value: Any, col_name: str, row_num: int) -> int:
    """페이지 번호 값 검증: 양의 정수만 허용"""
    try:
        page = int(value)
    except (TypeError, ValueError):
        raise ValueError(
            f"행 {row_num}: {col_name}은 양의 정수여야 합니다 (입력값: {value!r})"
        )
    if page <= 0:
        raise ValueError(
            f"행 {row_num}: {col_name}은 양의 정수여야 합니다 (입력값: {value!r})"
        )
    return page


def _normalize_str(value: Any) -> Optional[str]:
    """빈 문자열/공백만 있는 값을 None으로 정규화"""
    if value is None:
        return None
    if isinstance(value, str) and not value.strip():
        return None
    return str(value)


def _validate_context_page_pairs(
    pairs: List[tuple],
    row_num: int,
) -> None:
    """context/page 쌍 검증 (context_1/page_1은 필수)"""
    for ctx_val, page_val, n in pairs:
        ctx_col = f"reference_context_{n}"
        page_col = f"reference_page_{n}"

        # context_1과 page_1은 항상 필수
        if n == 1:
            if ctx_val is None:
                raise ValueError(
                    f"행 {row_num}: {ctx_col}은 필수입니다"
                )
            if page_val is None:
                raise ValueError(
                    f"행 {row_num}: {page_col}은 필수입니다"
                )

        if ctx_val is not None and page_val is None:
            raise ValueError(
                f"행 {row_num}: {ctx_col}이 존재하면 {page_col}도 필수입니다"
            )
        if ctx_val is None and page_val is not None:
            raise ValueError(
                f"행 {row_num}: {page_col}만 있고 {ctx_col}이 없습니다"
            )
        if page_val is not None:
            _validate_page_value(page_val, page_col, row_num)


def parse_excel_dataset(file: BinaryIO) -> List[RagasDatasetItem]:
    """
    Excel 골든 데이터셋을 파싱하여 RagasDatasetItem 리스트로 변환

    Args:
        file: Excel 파일 (BytesIO 또는 파일 객체)

    Returns:
        파싱된 데이터셋 아이템 리스트

    Raises:
        ValueError: 필수 열이 누락된 경우
    """
    wb = openpyxl.load_workbook(file, read_only=True)
    try:
        ws = wb["dataset"]

        # 헤더 읽기
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

        # 필수 열 검증
        missing = REQUIRED_COLUMNS - set(headers)
        if missing:
            raise ValueError(f"필수 열 누락: {', '.join(sorted(missing))}")

        # 열 인덱스 매핑
        col_idx = {name: idx for idx, name in enumerate(headers)}

        items: List[RagasDatasetItem] = []
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if row[col_idx["id"]] is None:
                continue

            # context/page 쌍 추출 (빈 문자열 → None 정규화)
            ctx_page_pairs = []
            for n in range(1, 4):
                ctx_col = f"reference_context_{n}"
                page_col = f"reference_page_{n}"
                ctx_val = _normalize_str(row[col_idx[ctx_col]] if ctx_col in col_idx else None)
                page_val = row[col_idx[page_col]]

                ctx_page_pairs.append((ctx_val, page_val, n))

            # 검증
            _validate_context_page_pairs(ctx_page_pairs, row_num)

            item = RagasDatasetItem.from_excel_row(
                id=row[col_idx["id"]],
                user_input=row[col_idx["user_input"]],
                category=row[col_idx["category"]],
                reference_context_1=ctx_page_pairs[0][0],
                reference_page_1=ctx_page_pairs[0][1],
                reference_context_2=ctx_page_pairs[1][0],
                reference_page_2=ctx_page_pairs[1][1],
                reference_context_3=ctx_page_pairs[2][0],
                reference_page_3=ctx_page_pairs[2][1],
                source_document=row[col_idx["source_document"]],
                response=row[col_idx["response"]] if "response" in col_idx else None,
            )
            items.append(item)

        return items
    finally:
        wb.close()


def validate_dataset_bytes(file_bytes: bytes) -> int:
    """업로드된 데이터셋 Excel을 즉시 검증한다 (레코드 생성 전 사용).

    parse_excel_dataset를 재사용해 시트명/필수 컬럼/행 형식을 검증하며,
    형식 오류 시 사용자에게 노출 가능한 메시지를 담은 ValueError를 발생시킨다.

    참고: response 컬럼은 업로드 시점에 비어 있어도 된다. generation/all 모드에서는
    평가 시작 시 에이전트(LLM) 응답으로 response를 채우므로, 업로드 검증에서는
    response 존재 여부를 강제하지 않는다.

    Args:
        file_bytes: 업로드된 Excel 파일 바이트

    Returns:
        int: 유효한 데이터셋 행 수

    Raises:
        ValueError: 시트/컬럼/행 형식 오류 또는 빈 데이터셋
    """
    try:
        items = parse_excel_dataset(io.BytesIO(file_bytes))
    except KeyError:
        # wb["dataset"] 미존재
        raise ValueError(
            "Excel에 'dataset' 시트가 없습니다. 시트명을 'dataset'으로 만들어 주세요."
        )
    except ValueError:
        # parse_excel_dataset이 만든 메시지(필수 열 누락 / 행 N: ...) 그대로 전달
        raise
    except Exception as exc:
        raise ValueError(f"Excel 파일을 읽을 수 없습니다: {exc}")

    if not items:
        raise ValueError("데이터셋에 유효한 행이 없습니다. (id가 채워진 행 필요)")

    return len(items)


def build_ragas_samples(
    user_input: str,
    search_results: List[Dict[str, Any]],
    reference_contexts: List[str],
    response: Optional[str] = None,
) -> SingleTurnSample:
    """
    검색 결과와 정답 근거를 RAGAS SingleTurnSample로 변환

    Args:
        user_input: 검색 질의
        search_results: HybridSearchService 검색 결과
        reference_contexts: 골든 데이터셋의 정답 근거 리스트
        response: LLM 생성 답변 (generation/all 모드에서 사용)

    Returns:
        RAGAS 평가용 SingleTurnSample
    """
    retrieved = [r["parsed_text"] for r in search_results]
    return SingleTurnSample(
        user_input=user_input,
        retrieved_contexts=retrieved,
        reference_contexts=reference_contexts,
        reference="\n".join(reference_contexts),
        response=response,
    )


def run_ragas_evaluation(
    samples: List[SingleTurnSample],
    max_workers: int = 16,
    eval_mode: str = "retrieval",
    llm_model: str = "gpt-4o",
) -> Dict[str, List[Optional[float]]]:
    """
    RAGAS 평가 실행 (eval_mode에 따른 지표 선택)

    Args:
        samples: SingleTurnSample 리스트
        max_workers: 병렬 실행 worker 수
        eval_mode: 평가 모드 (retrieval / generation / all)
        llm_model: 평가에 사용할 LLM 모델명

    Returns:
        메트릭별 개별 점수 딕셔너리
    """
    logger.info(
        f"RAGAS 평가 시작: {len(samples)}건, mode={eval_mode}, "
        f"model={llm_model}, max_workers={max_workers}"
    )

    # 평가용 AsyncOpenAI 클라이언트는 평가마다 생성하므로 반드시 닫아 httpx 누수를 방지한다.
    ragas_client = AsyncOpenAI()
    try:
        ragas_llm = llm_factory(llm_model, client=ragas_client)

        # RAGAS 0.4.3 버그 우회: gpt-5.4-mini 등 소수점 버전 모델 감지 실패
        if needs_max_completion_tokens(llm_model) and "max_tokens" in ragas_llm.model_args:
            ragas_llm.model_args["max_completion_tokens"] = ragas_llm.model_args.pop("max_tokens")
            logger.info(f"[RAGAS] max_tokens → max_completion_tokens 변환 적용 (model={llm_model})")

        metrics = _build_metrics(eval_mode, ragas_llm)

        dataset = EvaluationDataset(samples=samples)
        run_config = RunConfig(max_workers=max_workers, timeout=180)

        results = evaluate(dataset=dataset, metrics=metrics, run_config=run_config)

        scores = _extract_scores(results, eval_mode)

        logger.info(f"RAGAS 평가 완료: {', '.join(f'{k} 평균={_safe_mean(v):.4f}' for k, v in scores.items())}")

        return scores
    finally:
        _close_async_client(ragas_client)


def _build_metrics(eval_mode: str, ragas_llm: Any) -> list:
    """eval_mode에 따라 RAGAS 메트릭 리스트 생성 (호출마다 새 인스턴스)"""
    metrics = []

    if eval_mode in ("retrieval", "all"):
        metrics.append(ContextPrecision(llm=ragas_llm))
        metrics.append(ContextRecall(llm=ragas_llm))

    if eval_mode in ("generation", "all"):
        metrics.append(Faithfulness(llm=ragas_llm))
        metrics.append(AnswerRelevancy(
            llm=ragas_llm,
            embeddings=embedding_factory("text-embedding-ada-002"),
        ))

    return metrics


def _close_async_client(client: AsyncOpenAI) -> None:
    """sync 컨텍스트에서 AsyncOpenAI(httpx) 클라이언트를 안전하게 종료한다.

    평가마다 새로 생성한 클라이언트를 닫지 않으면 httpx 커넥션 풀이 누수된다.
    evaluate() 완료 후(성공/실패 무관) 호출되며, 닫기 실패는 로깅만 하고 무시한다.
    """
    try:
        asyncio.run(client.close())
    except Exception as exc:
        logger.warning(f"[RAGAS] 평가 클라이언트 종료 실패(무시): {exc}")


def _sanitize_score(value: Any) -> Optional[float]:
    """NaN/Inf → None 변환 (PostgreSQL JSONB 호환)"""
    if value is None:
        return None
    if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
        return None
    return float(value)


def _extract_scores(
    results: Any,
    eval_mode: str,
) -> Dict[str, List[Optional[float]]]:
    """evaluate() 결과에서 eval_mode에 해당하는 점수를 추출 (NaN → None)"""
    scores: Dict[str, List[Optional[float]]] = {}

    if eval_mode in ("retrieval", "all"):
        scores["context_precision"] = [_sanitize_score(v) for v in results["context_precision"]]
        scores["context_recall"] = [_sanitize_score(v) for v in results["context_recall"]]

    if eval_mode in ("generation", "all"):
        scores["faithfulness"] = [_sanitize_score(v) for v in results["faithfulness"]]
        scores["answer_relevancy"] = [_sanitize_score(v) for v in results["answer_relevancy"]]

    return scores


def _safe_mean(values: List[Optional[float]]) -> float:
    """None을 제외한 평균 계산"""
    valid = [v for v in values if v is not None]
    return sum(valid) / len(valid) if valid else 0.0


