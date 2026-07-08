import asyncio
import logging
import shutil
import tempfile
import os
from pathlib import Path
from typing import List, Dict
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell
import xlrd
from datetime import datetime, date

from app.parser.base import ParserInterface
from app.parser.utils.excel_chunker import ExcelTableChunker

logger = logging.getLogger(__name__)


class ExcelParser(ParserInterface):
    """
    Excel 파일 (.xlsx, .xls) 파서

    시트별 데이터 추출, 셀 데이터 타입 처리, 표 형태 데이터 구조화를 지원합니다.
    """

    def __init__(self):
        self.supported_extensions = [".xlsx", ".xls"]
        self.chunker = ExcelTableChunker()

    async def parsing(
        self, file_path: str, filename: str = None
    ) -> List[Dict]:
        """
        Excel 파일을 파싱하여 시트별 데이터를 추출합니다.

        Args:
            file_path: 파일 경로 (임시 파일 경로일 수 있음)
            filename: 원본 파일 이름 (확장자 확인용)

        Returns:
            List[Dict]: 시트별 파싱 결과 목록
        """
        try:
            logger.info(f"Excel 파일 파싱 시작: file_path={file_path}, filename={filename}")

            # 파일 확장자 확인 - filename에서 먼저 확인
            if filename:
                file_ext = Path(filename).suffix.lower()
            else:
                file_ext = Path(file_path).suffix.lower()

            if not file_ext:
                raise ValueError("파일 확장자를 확인할 수 없습니다")

            if file_ext not in self.supported_extensions:
                raise ValueError(f"지원되지 않는 파일 형식: {file_ext}")

            # 비동기 파싱 실행 (확장자 정보 전달)
            result = await self._parse_excel_async(file_path, file_ext)

            logger.info(f"Excel 파일 파싱 완료: {len(result)}개 시트 처리됨")

            return result

        except Exception as e:
            logger.error(f"Excel 파싱 중 오류 발생: {e}")
            raise

    async def _parse_excel_async(self, file_path: str, file_ext: str) -> List[Dict]:
        """Excel 파일 비동기 파싱"""
        return await asyncio.to_thread(self._parse_excel_sync, file_path, file_ext)

    def _parse_excel_sync(self, file_path: str, file_ext: str) -> List[Dict]:
        """Excel 파일 동기 파싱"""
        try:
            # 전달받은 확장자에 따라 적절한 파싱 방법 선택
            if file_ext == ".xlsx":
                return self._parse_xlsx(file_path)
            elif file_ext == ".xls":
                return self._parse_xls(file_path)
            else:
                raise ValueError(f"지원되지 않는 파일 형식: {file_ext}")

        except Exception as e:
            logger.error(f"Excel 파일 동기 파싱 중 오류: {e}")
            raise

    def _parse_xlsx(self, file_path: str) -> List[Dict]:
        """XLSX 파일 파싱 (OpenPyXL 사용)"""
        result_list = []
        temp_file_with_ext = None

        try:
            # 임시 파일에 확장자가 없으면 복사본 생성
            if not Path(file_path).suffix:
                # 확장자를 포함한 임시 파일 생성
                with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
                    temp_file_with_ext = tmp.name
                    shutil.copy2(file_path, temp_file_with_ext)
                    file_to_load = temp_file_with_ext
            else:
                file_to_load = file_path

            # 워크북 로드
            workbook = load_workbook(file_to_load, data_only=True)

            # 각 시트 처리
            for sheet_index, sheet_name in enumerate(workbook.sheetnames):
                sheet = workbook[sheet_name]

                # 시트 데이터 추출 (기본 모드로만 동작)
                sheet_data = self._extract_sheet_data(
                    sheet, sheet_name, sheet_index + 1
                )
                result = {
                    "page_number": sheet_index + 1,
                    "sheet_name": sheet_name,
                    "text": sheet_data["text"],
                    "tables": sheet_data["tables"],
                    "charts": [],
                    "images": [],
                    "metadata": {
                        "row_count": sheet_data["row_count"],
                        "column_count": sheet_data["column_count"],
                        "has_charts": False,
                        "has_images": False,
                    },
                }

                result_list.append(result)

        except Exception as e:
            logger.error(f"XLSX 파싱 중 오류: {e}")
            raise
        finally:
            if "workbook" in locals():
                workbook.close()
            # 임시 파일 정리
            if temp_file_with_ext and os.path.exists(temp_file_with_ext):
                try:
                    os.remove(temp_file_with_ext)
                except Exception:
                    pass

        return result_list

    def _parse_xls(self, file_path: str) -> List[Dict]:
        """XLS 파일 파싱 (xlrd 사용)"""
        result_list = []
        temp_file_with_ext = None

        try:
            # 임시 파일에 확장자가 없으면 복사본 생성
            if not Path(file_path).suffix:
                # 확장자를 포함한 임시 파일 생성
                with tempfile.NamedTemporaryFile(suffix='.xls', delete=False) as tmp:
                    temp_file_with_ext = tmp.name
                    shutil.copy2(file_path, temp_file_with_ext)
                    file_to_load = temp_file_with_ext
            else:
                file_to_load = file_path

            # 워크북 로드
            workbook = xlrd.open_workbook(filename=file_to_load)

            # 각 시트 처리
            for sheet_index in range(workbook.nsheets):
                sheet = workbook.sheet_by_index(sheet_index)
                sheet_name = workbook.sheet_names()[sheet_index]

                # 시트 데이터 추출
                sheet_data = self._extract_xls_sheet_data(
                    sheet, sheet_name, sheet_index + 1
                )

                # 결과 조합
                result = {
                    "page_number": sheet_index + 1,
                    "sheet_name": sheet_name,
                    "text": sheet_data["text"],
                    "tables": sheet_data["tables"],
                    "charts": [],  # XLS에서는 차트 추출 제한적
                    "images": [],  # XLS에서는 이미지 추출 제한적
                    "metadata": {
                        "row_count": sheet_data["row_count"],
                        "column_count": sheet_data["column_count"],
                        "has_charts": False,
                        "has_images": False,
                    },
                }

                result_list.append(result)

        except Exception as e:
            logger.error(f"XLS 파싱 중 오류: {e}")
            raise
        finally:
            # 임시 파일 정리
            if temp_file_with_ext and os.path.exists(temp_file_with_ext):
                try:
                    os.remove(temp_file_with_ext)
                except Exception:
                    pass

        return result_list

    def _extract_sheet_data(
        self, sheet: Worksheet, sheet_name: str, sheet_index: int
    ) -> Dict:
        """시트 데이터 추출 (병합 셀 처리)"""
        try:
            # 시트 크기 정보
            max_row = sheet.max_row
            max_col = sheet.max_column

            # 병합 셀 정보 추출
            merged_cells_info = self._get_merged_cells_info(sheet)

            # 데이터 추출 (병합 표시 안정 버전)
            table_data = []
            for row in range(1, max_row + 1):
                row_data = []
                has_meaningful_content = False

                for col in range(1, max_col + 1):
                    cell = sheet.cell(row=row, column=col)

                    # 병합 셀 처리
                    cell_value = self._extract_merged_cell_value(
                        cell, row, col, merged_cells_info
                    )

                    # 값 정규화
                    if cell_value:
                        cell_value = str(cell_value).strip()
                        # 빈 문자열이 아닌 실제 데이터만 의미있는 내용으로 간주
                        if cell_value and cell_value != "":
                            has_meaningful_content = True

                    row_data.append(cell_value if cell_value else "")

                # 의미있는 내용이 있는 행만 추가
                if has_meaningful_content:
                    # 일관된 테이블 구조를 위해 모든 컬럼 유지 (끝의 빈 값 제거 안함)
                    # 모든 행이 동일한 컬럼 수를 갖도록 보장
                    while len(row_data) < max_col:
                        row_data.append("")

                    table_data.append(row_data)

            # 테이블 데이터를 마크다운으로 변환
            markdown_tables = []
            if table_data:
                # 최소 2행 이상인 경우에만 테이블로 처리
                if len(table_data) >= 2:
                    markdown_table = self._convert_to_markdown_table_with_merge(
                        table_data, merged_cells_info
                    )
                    if markdown_table:
                        markdown_tables.append(markdown_table)
                else:
                    # 단일 행인 경우 일반 텍스트로 처리
                    single_row_text = " | ".join(
                        cell for cell in table_data[0] if cell.strip()
                    )
                    if single_row_text:
                        markdown_tables.append(f"**데이터:** {single_row_text}")

            # 텍스트 생성 (병합 정보 제거)
            text_content = f"# {sheet_name}\n\n"

            # 테이블 텍스트 추가
            if markdown_tables:
                text_content += "\n\n".join(markdown_tables)
            else:
                # 테이블이 없는 경우 대체 텍스트 처리
                if table_data:
                    text_lines = []
                    for row in table_data[:10]:
                        meaningful_data = [
                            cell.strip() for cell in row if cell and cell.strip()
                        ]
                        if meaningful_data:
                            text_lines.append(" | ".join(meaningful_data))

                    if text_lines:
                        text_content += "\n".join(text_lines)

            # 빈 시트 처리
            if not table_data:
                text_content += "*(빈 시트)*"

            return {
                "text": text_content,
                "tables": markdown_tables,
                "row_count": len(table_data),
                "column_count": max_col,
            }

        except Exception as e:
            logger.error(f"시트 데이터 추출 중 오류: {e}")
            raise

    def _get_merged_cells_info(self, sheet: Worksheet) -> Dict[tuple, Dict]:
        """병합 셀 정보 추출"""
        merged_info = {}

        try:
            for merged_range in sheet.merged_cells.ranges:
                # 병합 범위 정보
                min_row, min_col = merged_range.min_row, merged_range.min_col
                max_row, max_col = merged_range.max_row, merged_range.max_col

                # 병합된 셀의 실제 값 (첫 번째 셀에서 가져옴)
                master_cell = sheet.cell(row=min_row, column=min_col)
                master_value = self._extract_cell_value(master_cell)

                # 병합 범위의 모든 셀에 대해 정보 저장
                for row in range(min_row, max_row + 1):
                    for col in range(min_col, max_col + 1):
                        merged_info[(row, col)] = {
                            "value": master_value,
                            "is_master": (row == min_row and col == min_col),
                            "range": f"{min_row},{min_col}:{max_row},{max_col}",
                            "size": f"{max_row - min_row + 1}x{max_col - min_col + 1}",
                        }

        except Exception as e:
            logger.warning(f"병합 셀 정보 추출 중 오류: {e}")

        return merged_info

    def _extract_merged_cell_value(
        self, cell: Cell, row: int, col: int, merged_info: Dict[tuple, Dict]
    ) -> str:
        """병합 셀을 고려한 셀 값 처리 (빈 문자열 버전)"""
        try:
            # 병합 셀인지 확인
            if (row, col) in merged_info:
                merge_data = merged_info[(row, col)]
                # 마스터 셀인 경우에만 값 반환
                if merge_data["is_master"]:
                    value = merge_data["value"]
                    return str(value).strip() if value else ""
                else:
                    # 서브 셀은 빈 문자열로 처리 (Markdown 구분선 충돌 방지)
                    return ""
            else:
                # 일반 셀 처리
                cell_value = self._extract_cell_value(cell)
                return str(cell_value).strip() if cell_value else ""

        except Exception as e:
            logger.warning(f"병합 셀 값 처리 중 오류: {e}")
            # 오류 발생 시 기본 셀 값 처리
            cell_value = self._extract_cell_value(cell)
            return str(cell_value).strip() if cell_value else ""

    def _preprocess_table_data(self, data: List[List[str]]) -> List[List[str]]:
        """테이블 데이터 공통 전처리 (빈 열 제거, 대시 행 제거)"""
        if not data:
            return []

        # 1단계: 첫 번째 빈 열 제거 (인덱스용 빈 열 제거)
        if data and all(row and len(row) > 0 and row[0].strip() == "" for row in data):
            data = [row[1:] if len(row) > 1 else [] for row in data]
            data = [row for row in data if row]

        if not data:
            return []

        # 2단계: 순수 대시("-")로만 구성된 행 완전 제거
        filtered_data = []
        for row in data:
            if row:
                cleaned_cells = [cell.strip() for cell in row if cell is not None]
                if not all(cell == "-" or cell == "" for cell in cleaned_cells):
                    filtered_data.append(row)

        return filtered_data if filtered_data else []

    def _normalize_table_data(self, data: List[List[str]]) -> List[List[str]]:
        """테이블 데이터 정규화 (이스케이프 처리, 컬럼 수 통일)"""
        if not data:
            return []

        # 최대 컬럼 수 확인
        max_cols = max(len(row) for row in data) if data else 0
        if max_cols == 0:
            return []

        # 데이터 정규화 + 이스케이프 처리
        normalized_data = []
        for row in data:
            normalized_row = []
            for i in range(max_cols):
                if i < len(row):
                    cell_value = str(row[i]).strip() if row[i] else ""
                    # 특수 문자 이스케이프 (대시 포함)
                    cell_value = (
                        cell_value.replace("|", "\\|")
                        .replace("\n", " ")
                        .replace("\r", " ")
                        .replace("-", "\\-")
                    )
                    normalized_row.append(cell_value)
                else:
                    normalized_row.append("")

            # 빈 행이 아닌 경우에만 추가
            if any(cell.strip() for cell in normalized_row):
                normalized_data.append(normalized_row)

        return normalized_data

    def _optimize_table_columns(
        self, normalized_data: List[List[str]]
    ) -> tuple[List[List[str]], int]:
        """테이블 열 수 최적화 (중간 활성화 컬럼 고려)"""
        if not normalized_data:
            return [], 0

        # 현재 최대 컬럼 수
        max_cols = max(len(row) for row in normalized_data) if normalized_data else 0

        # 각 컬럼별 사용 빈도 계산 (중간 활성화 컬럼 감지)
        column_usage = [0] * max_cols
        total_rows = len(normalized_data)

        for row in normalized_data:
            for i in range(len(row)):
                if i < max_cols and row[i] and row[i].strip():
                    column_usage[i] += 1

        # 의미있는 컬럼 찾기 (사용률 5% 이상 또는 헤더에서 정의된 컬럼)
        meaningful_max_cols = max_cols
        usage_threshold = max(1, total_rows * 0.05)  # 최소 5% 사용률

        # 뒤에서부터 확인하여 의미없는 컬럼 제거
        for i in range(max_cols - 1, -1, -1):
            if column_usage[i] >= usage_threshold:
                meaningful_max_cols = i + 1
                break
            # 헤더 행(첫 번째 행)에서 정의된 컬럼은 보존
            if (
                normalized_data
                and len(normalized_data[0]) > i
                and normalized_data[0][i]
                and normalized_data[0][i].strip()
            ):
                meaningful_max_cols = i + 1
                break

        # 최소한 헤더에서 정의된 컬럼 수는 보장
        if normalized_data and len(normalized_data) > 0:
            header_cols = 0
            for i, cell in enumerate(normalized_data[0]):
                if cell and cell.strip():
                    header_cols = i + 1
            meaningful_max_cols = max(meaningful_max_cols, header_cols)

        # 열 수 조정이 필요한 경우
        if meaningful_max_cols > 0 and meaningful_max_cols < max_cols:
            # 모든 행을 의미있는 길이로 조정
            optimized_data = []
            for row in normalized_data:
                trimmed_row = row[:meaningful_max_cols]
                # 부족한 경우 빈 문자열로 패딩
                while len(trimmed_row) < meaningful_max_cols:
                    trimmed_row.append("")
                optimized_data.append(trimmed_row)
            return optimized_data, meaningful_max_cols

        return normalized_data, max_cols

    def _generate_table_header(
        self, normalized_data: List[List[str]], max_cols: int
    ) -> List[str]:
        """테이블 헤더 생성 및 처리"""
        if not normalized_data:
            return [f"컬럼{i+1}" for i in range(max_cols)]

        header = normalized_data[0] if normalized_data else []

        # 헤더가 없거나 모두 비어있는 경우 기본 헤더 생성
        if not header or all(not str(cell).strip() for cell in header):
            return [f"컬럼{i+1}" for i in range(max_cols)]

        # 헤더 길이를 max_cols에 맞춤
        header = header[:max_cols]
        while len(header) < max_cols:
            header.append(f"컬럼{len(header)+1}")

        return header

    def _extract_xls_sheet_data(self, sheet, sheet_name: str, sheet_index: int) -> Dict:
        """XLS 시트 데이터 추출"""
        try:
            # 시트 크기 정보
            max_row = sheet.nrows
            max_col = sheet.ncols

            # 데이터 추출
            table_data = []

            for row in range(max_row):
                row_data = []
                for col in range(max_col):
                    cell = sheet.cell(row, col)

                    # 셀 값 처리
                    cell_value = self._extract_xls_cell_value(cell)
                    row_data.append(cell_value)

                # 빈 행 제거
                if any(cell for cell in row_data):
                    table_data.append(row_data)

            # 테이블 데이터를 마크다운으로 변환 (4단계 전처리 자동 적용)
            markdown_tables = []
            if table_data:
                markdown_table = self._convert_to_markdown_table(table_data)
                if markdown_table:
                    markdown_tables.append(markdown_table)

            # 텍스트 생성
            text_content = f"# {sheet_name}\n\n"

            # 테이블 텍스트 추가
            if markdown_tables:
                text_content += "\n\n".join(markdown_tables)

            # 빈 시트 처리
            if not table_data:
                text_content += "*(빈 시트)*"

            return {
                "text": text_content,
                "tables": markdown_tables,
                "row_count": len(table_data),
                "column_count": max_col,
            }

        except Exception as e:
            logger.error(f"XLS 시트 데이터 추출 중 오류: {e}")
            raise

    def _extract_cell_value(self, cell: Cell) -> str:
        """셀 값 처리"""
        try:
            if cell.value is None:
                return ""

            # 데이터 타입에 따른 처리
            if cell.data_type == "n":  # 숫자
                if isinstance(cell.value, (int, float)):
                    # 정수인 경우 소수점 제거
                    if cell.value == int(cell.value):
                        return str(int(cell.value))
                    else:
                        return str(cell.value)
                else:
                    return str(cell.value)

            elif cell.data_type == "d":  # 날짜
                if isinstance(cell.value, datetime):
                    return cell.value.strftime("%Y-%m-%d %H:%M:%S")
                elif isinstance(cell.value, date):
                    return cell.value.strftime("%Y-%m-%d")
                else:
                    return str(cell.value)

            elif cell.data_type == "b":  # 불린
                return "TRUE" if cell.value else "FALSE"

            elif cell.data_type == "s":  # 문자열
                return str(cell.value).strip()

            elif cell.data_type == "f":  # 수식
                # 수식 결과 반환
                if cell.value is not None:
                    return str(cell.value)
                else:
                    return ""

            else:
                return str(cell.value).strip()

        except Exception as e:
            logger.warning(f"셀 값 처리 중 오류: {e}")
            return str(cell.value) if cell.value is not None else ""

    def _extract_xls_cell_value(self, cell) -> str:
        """XLS 셀 값 처리"""
        try:
            if cell.value is None:
                return ""

            # 셀 타입에 따른 처리
            if cell.ctype == xlrd.XL_CELL_EMPTY:
                return ""
            elif cell.ctype == xlrd.XL_CELL_TEXT:
                return str(cell.value).strip()
            elif cell.ctype == xlrd.XL_CELL_NUMBER:
                # 정수인 경우 소수점 제거
                if cell.value == int(cell.value):
                    return str(int(cell.value))
                else:
                    return str(cell.value)
            elif cell.ctype == xlrd.XL_CELL_DATE:
                # 날짜 처리
                date_tuple = xlrd.xldate_as_tuple(cell.value, 0)
                if date_tuple:
                    date_obj = datetime(*date_tuple)
                    return date_obj.strftime("%Y-%m-%d %H:%M:%S")
                else:
                    return str(cell.value)
            elif cell.ctype == xlrd.XL_CELL_BOOLEAN:
                return "TRUE" if cell.value else "FALSE"
            else:
                return str(cell.value).strip()

        except Exception as e:
            logger.warning(f"XLS 셀 값 처리 중 오류: {e}")
            return str(cell.value) if cell.value is not None else ""

    def _convert_to_markdown_table(self, data: List[List[str]]) -> str:
        """데이터를 마크다운 테이블로 변환 (최적화된 버전)"""
        try:
            if not data:
                return ""

            # 공통 전처리 함수 사용
            data = self._preprocess_table_data(data)
            if not data:
                return ""

            # 데이터 정규화
            normalized_data = self._normalize_table_data(data)
            if not normalized_data:
                return ""

            # 열 수 최적화
            optimized_data, max_cols = self._optimize_table_columns(normalized_data)
            if not optimized_data:
                return ""

            # 헤더 생성
            header = self._generate_table_header(optimized_data, max_cols)

            # 마크다운 테이블 생성
            lines = []

            # 헤더 행
            lines.append("| " + " | ".join(header) + " |")

            # 구분선
            lines.append("| " + " | ".join(["---"] * len(header)) + " |")

            # 데이터 행들
            for row in optimized_data[1:]:
                # 행 길이를 헤더 길이에 맞춤
                padded_row = list(row[: len(header)])
                while len(padded_row) < len(header):
                    padded_row.append("")
                lines.append("| " + " | ".join(padded_row) + " |")

            return "\n".join(lines)

        except Exception as e:
            logger.error(f"마크다운 테이블 변환 중 오류: {e}")
            # 안전한 폴백 처리
            try:
                if data:
                    cols = max(len(row) for row in data) if data else 1
                    fallback_lines = []

                    # 폴백 헤더
                    fallback_lines.append(
                        "| " + " | ".join([f"컬럼{i+1}" for i in range(cols)]) + " |"
                    )
                    fallback_lines.append("| " + " | ".join(["---"] * cols) + " |")

                    # 폴백 데이터 (최대 20행)
                    for row in data[:20]:
                        padded_row = list(row)
                        while len(padded_row) < cols:
                            padded_row.append("")
                        safe_row = [
                            (
                                str(cell).replace("|", "\\|").replace("-", "\\-")
                                if cell
                                else ""
                            )
                            for cell in padded_row[:cols]
                        ]
                        fallback_lines.append("| " + " | ".join(safe_row) + " |")

                    return "\n".join(fallback_lines)
                else:
                    return "빈 테이블"
            except Exception:
                return "테이블 변환 실패"

    def _convert_to_markdown_table_with_merge(
        self, data: List[List[str]], merged_info: Dict[tuple, Dict]
    ) -> str:
        """병합 정보를 포함한 마크다운 테이블 변환 (최적화된 버전)"""
        try:
            if not data:
                return ""

            # 공통 전처리 함수 사용
            data = self._preprocess_table_data(data)
            if not data:
                return ""

            # 데이터 정규화
            normalized_data = self._normalize_table_data(data)
            if not normalized_data:
                return ""

            # 열 수 최적화
            optimized_data, max_cols = self._optimize_table_columns(normalized_data)
            if not optimized_data:
                return ""

            # 헤더 생성
            header = self._generate_table_header(optimized_data, max_cols)

            # 마크다운 테이블 생성
            lines = []

            # 헤더 행
            lines.append("| " + " | ".join(header) + " |")

            # 구분선
            lines.append("| " + " | ".join(["---"] * len(header)) + " |")

            # 데이터 행들
            for row in optimized_data[1:]:
                # 행 길이를 헤더 길이에 정확히 맞춤
                padded_row = list(row[: len(header)])
                while len(padded_row) < len(header):
                    padded_row.append("")
                lines.append("| " + " | ".join(padded_row) + " |")

            return "\n".join(lines)

        except Exception as e:
            logger.error(f"병합 마크다운 테이블 변환 중 오류: {e}")
            # 안전한 폴백 처리
            try:
                if data:
                    cols = max(len(row) for row in data) if data else 1
                    fallback_lines = []

                    # 폴백 헤더
                    fallback_lines.append(
                        "| " + " | ".join([f"컬럼{i+1}" for i in range(cols)]) + " |"
                    )
                    fallback_lines.append("| " + " | ".join(["---"] * cols) + " |")

                    # 폴백 데이터 (최대 20행)
                    for row in data[:20]:
                        padded_row = list(row)
                        while len(padded_row) < cols:
                            padded_row.append("")
                        safe_row = [
                            (
                                str(cell).replace("|", "\\|").replace("-", "\\-")
                                if cell
                                else ""
                            )
                            for cell in padded_row[:cols]
                        ]
                        fallback_lines.append("| " + " | ".join(safe_row) + " |")

                    return "\n".join(fallback_lines)
                else:
                    return "빈 테이블"
            except Exception:
                return "테이블 변환 실패"

    # ================================================================================
    # 청킹 관련 메서드들 (헤더 고정 방식)
    # ================================================================================

    def _build_chunked_result(
        self, data: List[List[str]], sheet_name: str, method: str, chunk_size: int
    ) -> Dict:
        """
        테이블 데이터를 헤더 고정 방식으로 청킹하여 결과 구성
        """
        try:
            logger.info(f"청킹 시작: {sheet_name}, 방법={method}, 크기={chunk_size}")

            if not data:
                return self._create_single_chunk_result([], sheet_name)

            # 전처리
            preprocessed_data = self._preprocess_table_data(data.copy())
            if not preprocessed_data:
                return self._create_single_chunk_result([], sheet_name)

            normalized_data = self._normalize_table_data(preprocessed_data)
            if not normalized_data:
                return self._create_single_chunk_result([], sheet_name)

            # 청킹 방식 결정
            format_type = "structured" if method == "chunked_structured" else "markdown"

            # 새 청킹 유틸리티 사용
            chunks = self.chunker.chunk_table_data(
                data=normalized_data,
                sheet_name=sheet_name,
                chunk_size=chunk_size,
                format_type=format_type,
            )

            logger.info(f"청킹 완료: {len(chunks)}개 청크 생성")

            return {
                "chunks": chunks,
                "total_chunks": len(chunks),
                "sheet_name": sheet_name,
                "original_row_count": len(data),
                "processed_row_count": (
                    len(normalized_data) - 1 if len(normalized_data) > 1 else 0
                ),
            }

        except Exception as e:
            logger.error(f"청킹 결과 구성 중 오류: {e}")
            return self._create_single_chunk_result(data, sheet_name)

    def _create_single_chunk_result(
        self, data: List[List[str]], sheet_name: str
    ) -> Dict:
        """단일 청크 결과 생성 (데이터가 적거나 오류인 경우)"""
        if not data:
            content = f"# {sheet_name}\n\n*(빈 시트)*"
        elif len(data) == 1:
            # 헤더만 있는 경우
            headers = data[0]
            content = f"# {sheet_name}\n\n**컬럼명**: " + ", ".join(headers)
        else:
            # 정상적인 경우지만 크기가 작음
            content = self._build_markdown_chunk(data[0], data[1:], sheet_name, 1)

        chunks = [
            {
                "page_number": 1,
                "chunk_index": 0,
                "text": content,
                "row_count": len(data) - 1 if len(data) > 1 else 0,
                "type": "single",
            }
        ]

        return {
            "chunks": chunks,
            "total_chunks": 1,
            "sheet_name": sheet_name,
            "original_row_count": len(data),
            "processed_row_count": len(data) - 1 if len(data) > 1 else 0,
        }
